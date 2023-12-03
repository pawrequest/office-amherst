import os
import tempfile
from pathlib import Path

import pytest
from openpyxl.reader.excel import load_workbook
from pandas import Series

from office_am.assets import AssetContext, AssetManager
from office_am.office_tools.excel import get_rows
from office_am.dflt import DFLT_CONST


def top_three_rows(in_wb, out_wb):
    in_rows = get_rows(in_wb)
    out_rows = get_rows(out_wb)
    assert in_rows == out_rows, f"Rows do not match: {in_rows} != {out_rows}"


@pytest.fixture
def manager_context_fxt():
    fd, out_file = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    out_file = Path(out_file)
    with AssetContext(out_file=out_file) as context:
        yield context


@pytest.fixture
def am_fxt(manager_context_fxt) -> AssetManager:
    am = manager_context_fxt
    yield am


@pytest.fixture
def df_asset_fxt(am_fxt):
    asset: Series = am_fxt.row_from_id_or_serial('1111')
    yield asset



def styles_match(in_wb, out_wb) -> bool:
    in_styles = get_styles(in_wb)
    out_styles = get_styles(out_wb)
    return in_styles == out_styles


#
def get_styles(filename: Path, start=0, end=3):
    wb = load_workbook(filename)
    ws = wb.active
    styles = [[cell.style for cell in row] for row in ws.iter_rows(min_row=start + 1, max_row=end)]
    return styles


def test_id(df_asset_fxt):
    id_num = df_asset_fxt[DFLT_CONST.ID]
    assert id_num == '1111'


# def test_empty_serial(am_context_fxt):
#     radio = am_context_fxt.get_asset('3333')
#     assert radio.serial_number


def test_serial_num(am_fxt, df_asset_fxt):
    radio = am_fxt.row_from_id_or_serial('1111')
    assert radio[DFLT_CONST.SERIAL] == df_asset_fxt[DFLT_CONST.SERIAL]
    assert radio[DFLT_CONST.ID] == df_asset_fxt[DFLT_CONST.ID]


def test_df_asset_from_id(am_fxt, df_asset_fxt):
    radio = am_fxt.row_from_id_or_serial('1111')


def test_check_fw(am_fxt, df_asset_fxt):
    try:
        id_or = '1111'
        am_fxt.check_fw(id_or, fw_version='XXXX')
        am_fxt.check_fw(id_or)
    except Exception as e:
        raise e


def test_get_assets(am_fxt):
    nums = ['1111', '2222', '3333']
    assets = am_fxt.df_a.loc[am_fxt.df_a['Number'].isin(nums)]
    assert len(assets) == 3
    assert assets.iloc[1][DFLT_CONST.ID] == '2222'


def test_smth(am_fxt: AssetManager, df_asset_fxt):
    radio = df_asset_fxt
    rad = am_fxt.row_from_id_or_serial('1111')
    am_fxt.set_field_by_id_or_serial('1111', DFLT_CONST.FW, 'jjjjjjjj')
    assert am_fxt.field_from_id_or_serial('1111', DFLT_CONST.FW) == 'jjjjjjjj'
    ...
