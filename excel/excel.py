import numbers
import os
from typing import Union, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

pathy = Union[str, os.PathLike]


def df_is_numeric(col):
    try:
        pd.to_numeric(col)
        return True
    except:
        return False


def coerce_df_numtype(df: pd.DataFrame, col_header: str, data: str | int | float):
    if df_is_numeric(df[col_header]) and not isinstance(data, numbers.Number):
        if input(f"TypeMismatch : cast {col_header} ({df[col_header].dtype}) to {type(data)}?") != 'y':
            exit("Aborted.")
        df[col_header] = df[col_header].astype(type(data))


def edit_excel_batch(infile: pathy, outfile: pathy, sheet: str, header_i: int, id_header: str, value_data: Iterable,
                     value_header: str,
                     data_insert: str):
    df = pd.read_excel(infile, sheet_name=sheet, header=header_i)
    coerce_df_numtype(df, col_header=value_header, data=data_insert)

    skipped = []
    for data in value_data:
        if not set_data(df, id_data=data, id_header=id_header, value_header=value_header, value_data=data_insert):
            skipped.append(data)

    if skipped and input(f"Missed: {skipped}\nContinue?") != 'y': exit()

    write_excel(outfile, df, sheet)


def write_excel(outfile: pathy, df: pd.DataFrame, sheet: str):
    try:
        df.to_excel(outfile, sheet_name=sheet, index=False)
    except PermissionError:
        print("Close Excel and retry.")
    except Exception as e:
        print(f"Error: {e}")
    else:
        print("Saved.")
        return 0



def check_data(df: pd.DataFrame, id_data: str | int, id_header: str, value_header: str,
               value_data: str | int | float) -> bool:
    """ Returns True if value_data is in col value_header for row id_data."""
    row = df[df[id_header].astype(str) == str(id_data)]
    if len(row) != 1:
        raise ValueError()
    res = value_data == row[value_header].iloc[0]
    print(f'{value_data} is {"not" if not res else ""} in {value_header} for {id_data}')
    return res


def set_data(df: pd.DataFrame, id_data: str | int, id_header: str, value_header: str, value_data: str | int | float):
    coerce_df_numtype(df=df, col_header=value_header, data=value_data)
    index_to_set = df[df[id_header].astype(str) == str(id_data)].index
    if len(index_to_set) != 1:
        raise ValueError()
    df.at[index_to_set[0], value_header] = value_data
    return True


def get_data_from_excel(df:pd.DataFrame, id_data: str, id_header: str, value_header: str):
    """ Returns True if value_data is in col value_header for row id_data.
    :param df: DataFrame to search
    :param id_data: Data to search for
    :param id_header: Header for id of record to search
    :param value_header: Column name for data to return
    """
    row = df[df[id_header].str.upper() == id_data.upper()]
    if len(row) != 1:
        print(f"No or multiple results found for {id_data}")
        input("Press enter to continue...")
    actual_data = row[value_header].iloc[0]
    return actual_data


def get_matching(df, key_column, result_column, value):
    result_values = df.loc[df[key_column].astype(str) == value, result_column].values
    return result_values


def convert(df, value, key_column, result_column):
    result_values = get_matching(df=df, key_column=key_column, result_column=result_column, value=value)
    if result_values.size == 0 or pd.isna(result_values[0]):
        replacement_value = input(f"No result found for {value}. Enter new value: ")
        if replacement_value:
            df.loc[df[key_column].astype(str) == value, result_column] = replacement_value
            return replacement_value  # return the new value
        raise ValueError(f"No result found for {value}")
    if result_values.size != 1:
        raise ValueError(f"Multiple results found for {value}: {', '.join(map(str, result_values))}")
    return result_values[0]


def df_to_wb(workbook, sheet, df, header_row, out_file):
    wb = load_workbook(workbook)
    ws = wb[sheet]
    rows = dataframe_to_rows(df, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + header_row, column=c_idx, value=value)
    wb.save(out_file)
