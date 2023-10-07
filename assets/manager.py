from dataclasses import dataclass
from enum import Enum
from pathlib import Path

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from excel.excel import check_data, set_data
from exceptions import id_exception_handler


def _an_id(id_or_serial):
    return len(id_or_serial) == 4


@dataclass
class Identity:
    def __init__(self, df, id_or_serial=None):
        self.df = df
        if id_or_serial is None:
            raise ValueError("Must provide either ID or serial number.")
        self.id_number = self._get_id(id_or_serial)
        self.serial_number = self._get_serial(id_or_serial)

    @id_exception_handler
    def _id_to_serial(self, radio_id):
        return convert(df=self.df, value=radio_id, key_column=DFLT.ID.value, result_column=DFLT.SERIAL.value)

    @id_exception_handler
    def _serial_to_id(self, serial):
        return convert(df=self.df, value=serial, key_column=DFLT.SERIAL.value, result_column=DFLT.ID.value)

    def _get_id(self, id_or_serial):
        return id_or_serial if _an_id(id_or_serial) else self._serial_to_id(id_or_serial)

    def _get_serial(self, id_or_serial):
        return id_or_serial if not _an_id(id_or_serial) else self._id_to_serial(id_or_serial)


class DFLT(Enum):
    SERIAL = 'Barcode'
    ID = 'Number'
    FW = 'FW'
    FW_VERSION = 'XXXX'
    WB = Path(__file__).resolve().with_name("assets_in.xlsx")
    OUTPUT = Path(__file__).resolve().with_name("assets_out.xlsx")
    SHEET = 'Sheet1'
    HEAD = 2


def get_matching(df, key_column, result_column, value):
    result_values = df.loc[df[key_column].astype(str) == value, result_column].values
    return result_values


def convert(df, value, key_column, result_column):
    result_values = get_matching(df=df, key_column=key_column, result_column=result_column, value=value)
    if result_values.size == 0 or pd.isna(result_values[0]):
        replacement_value = input(f"No result found for {value}. Enter new value: ")
        if replacement_value:
            df.loc[df[key_column].astype(str) == value, result_column] = replacement_value
            return replacement_value  # return the new value
        raise ValueError(f"No result found for {value}")
    if result_values.size != 1:
        raise ValueError(f"Multiple results found for {value}: {', '.join(map(str, result_values))}")
    return result_values[0]


class AssetManagerContext:
    def __init__(self, workbook=DFLT.WB.value, sheet=DFLT.SHEET.value, header_row=int(DFLT.HEAD.value), out_file=None):
        self.out_file = out_file or workbook
        self.wb = load_workbook(workbook)
        self.ws = self.wb[sheet]
        self.header_row = header_row
        self.df = pd.read_excel(workbook, sheet_name=sheet, header=header_row)

    def __enter__(self):
        self.asset_manager = AssetManager(self.df)
        return self.asset_manager

    def __exit__(self, exc_type, exc_val, exc_tb):
        rows = dataframe_to_rows(self.asset_manager.df, index=False, header=True)
        df_to_wb(wb=self.wb, ws=self.ws, rows=rows, header_row=self.header_row, out_file=self.out_file)


def df_to_wb(wb, ws, rows, header_row, out_file):
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + header_row, column=c_idx, value=value)
    wb.save(out_file)


class AssetManager:

    def __init__(self, df):
        self.df = df

    def set_fw(self, radio: Identity):
        radio_id = radio.id_number
        set_data(self.df, id_data=radio_id, id_header=DFLT.ID.value, value_header=DFLT.FW.value,
                 value_data=DFLT.FW_VERSION.value)
        return True

    def check_progged(self, id_to_check: Identity, fw_version=DFLT.FW_VERSION.value):
        return check_data(self.df, id_data=id_to_check.id_number, id_header=DFLT.ID.value, value_header=DFLT.FW.value,
                          value_data=fw_version)

#
#
# def get_function_to_call():
#     while True:
#         choice = input("Set or check? (s/c)").lower()
#         if choice in ('s', 'c'):
#             return set_progged if choice == 's' else check_progged
#         print("Invalid option. Use 's' for set or 'c' for check.")

#
# def handle_id(id_to_handle, action, df=None, serial_header=DFLT.SERIAL.value, id_header=DFLT.ID.value,
#               data_header=DFLT.FW.value, value=DFLT.FW_VERSION.value):
#     if not isinstance(id_to_handle, (str, int)):
#         raise TypeError("ID should be of type str or int")
#
#     if df is None:
#         df = get_excel()
#
#     col_to_use = id_header if len(str(id_to_handle)) == 4 else serial_header
#     return action(df, id_to_handle, col_to_use, data_header, value)
#
#
# def set_progged(df, id_to_handle, value_header, data_header, value, output_file=DFLT_OUTPUT_FILE):
#     set_data(df, id_data=id_to_handle, id_header=value_header, value_header=data_header, value_data=value)
#     df.to_excel(output_file, index=False)
#     print("Excel sheet successfully updated.")
#     return True
#
#
# def battery_test():
#     ...
#
#
# #
# # def set_fw_updated(df, id_data):
# #     set_data(df, id_data=id_data, id_header=value_header, value_header=data_header, value_data=value)
#
# def check_progged(df, id_to_check, col_to_check, data_header, value):
#     return check_data(df, id_data=id_to_check, id_header=col_to_check, value_header=data_header,
#                       value_data=value)
#
#
# def set_or_check():
#     function_to_call = get_function_to_call()
#     while True:
#         id_to_handle = input("Enter ID\n")
#         try:
#             handle_id(id_to_handle, function_to_call)
#         except ValueError:
#             print(f"No or multiple results found for {id_to_handle}.")
#         except PermissionError:
#             print("Close Excel and retry.")
#         except TypeError as e:
#             print(e)
#
#
# if __name__ == "__main__":
#     try:
#         logging.basicConfig(filename='assets.log', level=logging.DEBUG)
#         set_or_check()
#     except Exception as e:
#         logging.error(f"An error occurred: \n{e}")
#         input("Press Enter to exit.")

#
# @dataclass
# class Radio:
#     model: str
#     serial: str
#     id: str
#     fw: str


# class AssetManagerContextOLD:
#     def __init__(self, workbook=DFLT.WB.value, sheet=DFLT.SHEET.value, header_row=int(DFLT.HEAD.value), out_file=None):
#         self.out_file = out_file or workbook
#         self.workbook = workbook
#         self.sheet = sheet
#         self.header_row = header_row
#         self.df = pd.read_excel(self.workbook, sheet_name=self.sheet, header=self.header_row)
#
#     def __enter__(self):
#         # shutil.copyfile(self.workbook, self.out_file)
#
#         self.asset_manager = AssetManager(self.df)
#         return self.asset_manager
#
#     def __exit__(self, exc_type, exc_val, exc_tb):
#         self.asset_manager.df.to_excel(self.out_file, header=False, index=False, startrow=self.header_row + 1)
#
